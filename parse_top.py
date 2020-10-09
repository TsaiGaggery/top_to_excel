import os, sys, getopt
import csv
from os import path
import openpyxl
import re
from openpyxl.styles import PatternFill as PatternFill
from parse import compile
import xlsxwriter, re

def main(argv):
    inputfile = ''
    outfolder = ''

    try:
        opts, args = getopt.getopt(argv,"hi:o:", ["ifile=", "ofolder="])
    except getopt.GetoptError:
        print ("test.py -i <input file>  -o <output folder>")
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print ("test.py -i <input file> -o <output folder>")
            sys.exit()
        elif opt in ("-i", "--ifile"):
            if not path.exists(arg):
                print ("The input file %s does not exist!" % arg)
                sys.exit()
            inputfile = arg        
        elif opt in ("-o", "--ofolder"):
            if not path.exists(arg):
                print ("The output folder %s does not exist!" % arg)
                sys.exit()
            outfolder = arg

    print ("Input file is ", inputfile)
    print ("Output folder is ", outfolder)
    outfile = str(inputfile).split('/')[-1].split('.')[0] + '-output.xlsx'
    print ("Output file name is ", outfile)

    # create a output file
    workbook = xlsxwriter.Workbook(outfile)
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1})

    # define start from A1 cell
    start_row = 1
    current_column = 'A' 

    headings = {'time': 'A', \
                'up_time': 'B', \
                'users': 'C', \
                'load_avg': 'D', \
                'tasks': 'E', \
                'running_tasks': 'F', \
                'sleeping_tasks': 'G', \
                'stopped_tasks': 'H', \
                'zombie': 'I', \
                'cpu_usage_user': 'J', \
                'cpu_usage_sys': 'K', 
                'ni': 'L', \
                'id': 'M', \
                'wa': 'N', \
                'hi': 'O', \
                'si': 'P', \
                'st': 'Q', \
                'total_mem': 'R', \
                'free_mem': 'S', \
                'used_mem': 'T', \
                'buff_cache': 'U', \
                'swap_mem': 'V', \
                'free_swap': 'W', \
                'used_swap': 'X', \
                'available_mem': 'Y'}

    # write header file
    worksheet.write_row(current_column + str(start_row), list(headings.keys()), bold)
    start_row += 1
    num_cell_format = workbook.add_format()
    num_cell_format.set_num_format(2)
    time_cell_format = workbook.add_format()
    time_cell_format.set_num_format(21)

    try:
        fd = open(inputfile)
    except:
        print('Open file {} failed'.format(inputfile))
        sys.exit()

    for line in fd:
        if line.startswith('top'):
            try:   
                result=re.match('^top -\s(?P<now_time>.+?)\sup\s(?P<up_time>.+?),\s(?P<users_count>.+?)\susers,  load average:\s(?P<load_avg>.+?)$', line)
            except:
                print('Error while parsing the line {}'.format(line))
                sys.exit()            
            # fill the data to excel
            worksheet.write(current_column + str(start_row), result.group('now_time'), time_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), result.group('up_time'), time_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), int(result.group('users_count')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), result.group('load_avg'))
            current_column = chr(ord(current_column)+1)
        elif line.startswith('Tasks'):
            try:
                result=re.match('^Tasks:\s(?P<tasks>.+?)\stotal,\s(?P<running_tasks>.+?)\srunning,\s(?P<sleeping_tasks>.+?)\ssleeping,\s(?P<stopped_tasks>.+?)\sstopped,\s(?P<zombie>.+?)\szombie$', line)
            except:
                print('Error while parsing the line {}'.format(line))
                sys.exit()
            # fill the data to excel
            worksheet.write(current_column + str(start_row), int(result.group('tasks')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), int(result.group('running_tasks')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), int(result.group('sleeping_tasks')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), int(result.group('stopped_tasks')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), int(result.group('zombie')), num_cell_format)
            current_column = chr(ord(current_column)+1)
        elif line.startswith('%Cpu'):
            try:
                result=re.match('^%Cpu\\(s\\):\s(?P<cpu_usage_user>.+?)\sus,\s(?P<cpu_usage_sys>.+?)\ssy,\s(?P<ni>.+?)\sni,\s(?P<id>.+?)\sid,\s(?P<wa>.+?)\swa,\s(?P<hi>.+?)\shi,\s(?P<si>.+?)\ssi,\s(?P<st>.+?)\sst$', line)
            except:
                print('Error while parsing the line {}'.format(line))
                sys.exit()
            # fill the data to excel
            worksheet.write(current_column + str(start_row), float(result.group('cpu_usage_user')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('cpu_usage_sys')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('ni')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('id')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('wa')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('hi')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('si')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('st')), num_cell_format)
            current_column = chr(ord(current_column)+1)
        elif line.startswith('MiB Mem'):
            try:
                result=re.match('^MiB Mem :\s(?P<total_mem>.+?)\stotal,\s(?P<free_mem>.+?)\sfree,\s(?P<used_mem>.+?)\sused,\s(?P<buff_cache>.+?)\sbuff/cache$', line)
            except:
                print('Error while parsing the line {}'.format(line))
                sys.exit()

            # fill the data to excel
            worksheet.write(current_column + str(start_row), float(result.group('total_mem')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('free_mem')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('used_mem')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('buff_cache')), num_cell_format)
            current_column = chr(ord(current_column)+1)
        elif line.startswith('MiB Swap'):
            try:
                result=re.match('^MiB Swap:\s(?P<swap_mem>.+?)\stotal,\s(?P<free_swap>.+?)\sfree,\s(?P<used_swap>.+?)\sused\\.\s(?P<available_mem>.+?)\savail Mem $', line)
            except:
                print('Error while parsing the line {}'.format(line))
                sys.exit()

            # fill the data to excel
            worksheet.write(current_column + str(start_row), float(result.group('swap_mem')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('free_swap')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('used_swap')), num_cell_format)
            current_column = chr(ord(current_column)+1)
            worksheet.write(current_column + str(start_row), float(result.group('available_mem')), num_cell_format)
            current_column = chr(ord(current_column)+1)

            # Supposedly we reach the end of column, advance one row and re-position column back to A
            start_row += 1
            current_column = 'A'

    # Create a new chart object.
    chart = workbook.add_chart({'type': 'line'})

    # Add a series about CPU usage to the chart.
    chart.add_series({'name': 'cpu_usage_user', 'values': '=Sheet1!$' + headings['cpu_usage_user'] + '1:$'+ headings['cpu_usage_user']+str(start_row)})
    chart.add_series({'name': 'cpu_usage_sys', 'values': '=Sheet1!$' + headings['cpu_usage_sys'] + '1:$'+ headings['cpu_usage_sys']+str(start_row)})

    # Insert the chart into the worksheet.
    worksheet.insert_chart('C5', chart)   

    # Create a new chart object.
    chart = workbook.add_chart({'type': 'line'})

    # Add a series about Memory usage to the chart.
    chart.add_series({'name': 'free_mem', 'values': '=Sheet1!$' + headings['free_mem'] + '1:$'+ headings['free_mem']+str(start_row)})
    chart.add_series({'name': 'used_mem', 'values': '=Sheet1!$' + headings['used_mem'] + '1:$'+ headings['used_mem']+str(start_row)})
    chart.add_series({'name': 'buff_cache', 'values': '=Sheet1!$' + headings['buff_cache'] + '1:$'+ headings['buff_cache']+str(start_row)})
    chart.add_series({'name': 'available_mem', 'values': '=Sheet1!$' + headings['available_mem'] + '1:$'+ headings['available_mem']+str(start_row)})

    # Insert the chart into the worksheet.
    worksheet.insert_chart('K5', chart)   

    workbook.close()


if __name__ == "__main__":
    main(sys.argv[1:])

